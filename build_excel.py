from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPO_ROOT = Path(__file__).resolve().parent
DATA_DIR = REPO_ROOT / "data"
OUTPUT_PATH = REPO_ROOT / "excel" / "portfolio_dashboard.xlsx"


# Palette
DARK_NAVY = "1B2A4A"
HEADER_FONT_COLOR = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9EAF7"
LIGHT_GREEN = "DDF3D8"
GREEN_FILL = "C6EFCE"
YELLOW_FILL = "FFEB9C"
RED_FILL = "FFC7CE"
ORANGE_FILL = "F4B942"
DARK_GREEN_FILL = "2E7D32"
MID_GREEN_FILL = "A9D18E"
GRAY_FILL = "D9D9D9"
DARK_ORANGE = "C55A11"
RED_TAB = "C00000"
DARK_RED_TEXT = "9C0006"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

CARD_BORDER = Border(
    left=Side(style="medium", color="1B2A4A"),
    right=Side(style="medium", color="1B2A4A"),
    top=Side(style="medium", color="1B2A4A"),
    bottom=Side(style="medium", color="1B2A4A"),
)


def normalize_pct_value(value):
    if pd.isna(value):
        return value
    return value / 100 if value > 1.5 else value


def normalize_pct_series(series: pd.Series) -> pd.Series:
    return series.apply(normalize_pct_value)


def auto_fit_columns(ws, min_width=12, max_width=35):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col_cells:
            if type(cell).__name__ == "MergedCell":
                continue
            value = cell.value
            if value is None:
                continue
            if isinstance(value, pd.Timestamp):
                text = value.strftime("%b-%y")
            else:
                text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def style_table_header(ws, header_row=1):
    header_fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cells(ws, start_row, end_row, end_col):
    data_font = Font(name="Calibri", size=10)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.font = data_font
            if cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1):
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)


def apply_border_range(ws, min_row, max_row, min_col, max_col, border):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def style_custom_header(ws, row_idx, start_col, end_col):
    fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def build_cover_tab(wb: Workbook):
    ws = wb.create_sheet("Cover")
    ws.sheet_properties.tabColor = DARK_NAVY

    # Fill canvas
    bg_fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    for r in range(1, 35):
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = bg_fill

    ws.merge_cells("B4:I6")
    ws["B4"] = "Real Estate Portfolio Performance & Risk Dashboard"
    ws["B4"].font = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B8:I8")
    ws["B8"] = "Crystal View Capital / Osprey Management"
    ws["B8"].font = Font(name="Calibri", size=14, bold=False, color="D9D9D9")
    ws["B8"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B11:I11")
    ws["B11"] = "Portfolio Period: January – December 2024"
    ws["B11"].font = Font(name="Calibri", size=12, color="FFFFFF")
    ws["B11"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B13:I13")
    ws["B13"] = "12 Properties | 2 Asset Types | NV · AZ · NM · UT"
    ws["B13"].font = Font(name="Calibri", size=12, italic=True, color="B8C2D6")
    ws["B13"].alignment = Alignment(horizontal="center", vertical="center")

    for col in ("A", "J"):
        ws.column_dimensions[col].width = 4
    for col in ("B", "C", "D", "E", "F", "G", "H", "I"):
        ws.column_dimensions[col].width = 18
    ws.row_dimensions[4].height = 34
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 34


def build_properties_tab(wb: Workbook, properties_df: pd.DataFrame):
    ws = wb.create_sheet("Properties")
    ws.sheet_properties.tabColor = DARK_NAVY

    df = properties_df.rename(
        columns={
            "property_id": "Property ID",
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "city": "City",
            "state": "State",
            "total_units": "Total Units",
            "year_acquired": "Year Acquired",
            "purchase_price": "Purchase Price",
        }
    )

    write_dataframe(ws, df, start_row=1)
    style_table_header(ws)
    style_data_cells(ws, 2, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"

    for r in range(2, ws.max_row + 1):
        row_fill = PatternFill(fill_type="solid", fgColor=WHITE if r % 2 == 0 else LIGHT_GRAY)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = row_fill

        asset_cell = ws.cell(row=r, column=3)
        if asset_cell.value == "MHC":
            asset_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
        elif asset_cell.value == "Self-Storage":
            asset_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GREEN)

        ws.cell(row=r, column=8).number_format = "$#,##0.00"

    auto_fit_columns(ws)


def build_occupancy_tab(wb: Workbook, properties_df: pd.DataFrame, occupancy_df: pd.DataFrame):
    ws = wb.create_sheet("Occupancy")
    ws.sheet_properties.tabColor = "4F81BD"

    occ = occupancy_df.copy()
    occ["month"] = pd.to_datetime(occ["month"])
    occ["occupancy_rate"] = normalize_pct_series(occ["occupancy_rate"])

    occ = occ.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )

    def occupancy_flag(rate):
        if rate >= 0.93:
            return "On Target"
        if rate >= 0.88:
            return "Watch"
        return "Below Threshold"

    occ["Flag"] = occ["occupancy_rate"].apply(occupancy_flag)
    occ.sort_values(["month", "occupancy_rate"], ascending=[True, True], inplace=True)

    out = occ[
        [
            "property_id",
            "property_name",
            "asset_type",
            "month",
            "occupied_units",
            "total_units",
            "occupancy_rate",
            "Flag",
        ]
    ].rename(
        columns={
            "property_id": "Property ID",
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "month": "Month",
            "occupied_units": "Occupied Units",
            "total_units": "Total Units",
            "occupancy_rate": "Occupancy Rate",
        }
    )

    write_dataframe(ws, out, start_row=1)
    style_table_header(ws)
    style_data_cells(ws, 2, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "mmm-yy"
        ws.cell(row=r, column=7).number_format = "0.0%"

        flag_cell = ws.cell(row=r, column=8)
        if flag_cell.value == "On Target":
            flag_cell.fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
        elif flag_cell.value == "Watch":
            flag_cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
        else:
            flag_cell.fill = PatternFill(fill_type="solid", fgColor=RED_FILL)

    auto_fit_columns(ws)


def build_delinquency_tab(wb: Workbook, properties_df: pd.DataFrame, delinquency_df: pd.DataFrame):
    ws = wb.create_sheet("Delinquency")
    ws.sheet_properties.tabColor = "ED7D31"

    dlq = delinquency_df.copy()
    dlq["month"] = pd.to_datetime(dlq["month"])
    dlq["delinquency_rate"] = normalize_pct_series(dlq["delinquency_rate"])

    dlq = dlq.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )

    def risk_tier(rate):
        if rate < 0.05:
            return "Acceptable"
        if rate < 0.08:
            return "Monitor"
        if rate < 0.12:
            return "At Risk"
        return "Critical"

    dlq["Risk Tier"] = dlq["delinquency_rate"].apply(risk_tier)

    out = dlq[
        [
            "property_id",
            "property_name",
            "asset_type",
            "month",
            "rent_billed",
            "rent_collected",
            "past_due_amount",
            "delinquency_rate",
            "Risk Tier",
        ]
    ].rename(
        columns={
            "property_id": "Property ID",
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "month": "Month",
            "rent_billed": "Rent Billed",
            "rent_collected": "Rent Collected",
            "past_due_amount": "Past Due Amount",
            "delinquency_rate": "Delinquency Rate",
        }
    )

    write_dataframe(ws, out, start_row=1)
    style_table_header(ws)
    style_data_cells(ws, 2, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "mmm-yy"
        for c in (5, 6, 7):
            ws.cell(row=r, column=c).number_format = "$#,##0.00"
        ws.cell(row=r, column=8).number_format = "0.0%"

        tier_cell = ws.cell(row=r, column=9)
        if tier_cell.value == "Acceptable":
            tier_cell.fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
        elif tier_cell.value == "Monitor":
            tier_cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
        elif tier_cell.value == "At Risk":
            tier_cell.fill = PatternFill(fill_type="solid", fgColor=ORANGE_FILL)
        else:
            tier_cell.fill = PatternFill(fill_type="solid", fgColor=RED_FILL)

    auto_fit_columns(ws)


def build_financials_tab(wb: Workbook, properties_df: pd.DataFrame, financials_df: pd.DataFrame):
    ws = wb.create_sheet("Financials")
    ws.sheet_properties.tabColor = "70AD47"

    fin = financials_df.copy()
    fin["month"] = pd.to_datetime(fin["month"])
    fin["noi_margin"] = normalize_pct_series(fin["noi_margin"])
    fin["expense_ratio"] = normalize_pct_series(fin["expense_ratio"])

    fin = fin.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )

    out = fin[
        [
            "property_id",
            "property_name",
            "asset_type",
            "month",
            "total_revenue",
            "operating_expenses",
            "noi",
            "noi_margin",
            "expense_ratio",
        ]
    ].rename(
        columns={
            "property_id": "Property ID",
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "month": "Month",
            "total_revenue": "Total Revenue",
            "operating_expenses": "Operating Expenses",
            "noi": "NOI",
            "noi_margin": "NOI Margin",
            "expense_ratio": "Expense Ratio",
        }
    )

    write_dataframe(ws, out, start_row=1)
    style_table_header(ws)
    style_data_cells(ws, 2, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).number_format = "mmm-yy"
        for c in (5, 6, 7):
            ws.cell(row=r, column=c).number_format = "$#,##0.00"
        ws.cell(row=r, column=8).number_format = "0.0%"
        ws.cell(row=r, column=9).number_format = "0.0%"

        margin = ws.cell(row=r, column=8).value
        margin_cell = ws.cell(row=r, column=8)
        if margin >= 0.45:
            margin_cell.fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
        elif margin >= 0.35:
            margin_cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
        else:
            margin_cell.fill = PatternFill(fill_type="solid", fgColor=RED_FILL)

    auto_fit_columns(ws)


def build_market_comps_tab(wb: Workbook, properties_df: pd.DataFrame, market_df: pd.DataFrame):
    ws = wb.create_sheet("Market Comps")
    ws.sheet_properties.tabColor = "8064A2"

    mc = market_df.copy()
    mc["rent_upside_pct"] = normalize_pct_series(mc["rent_upside_pct"])
    mc["competitor_occupancy"] = normalize_pct_series(mc["competitor_occupancy"])
    mc["market_cap_rate"] = normalize_pct_series(mc["market_cap_rate"])

    mc = mc.merge(
        properties_df[["property_id", "property_name", "asset_type", "city", "state"]],
        on="property_id",
        how="left",
    )

    out = mc[
        [
            "property_name",
            "asset_type",
            "city",
            "state",
            "market_avg_rent",
            "property_avg_rent",
            "rent_upside_pct",
            "competitor_occupancy",
            "market_cap_rate",
            "opportunity_score",
        ]
    ].rename(
        columns={
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "city": "City",
            "state": "State",
            "market_avg_rent": "Market Avg Rent",
            "property_avg_rent": "Property Avg Rent",
            "rent_upside_pct": "Rent Upside %",
            "competitor_occupancy": "Competitor Occupancy",
            "market_cap_rate": "Market Cap Rate",
            "opportunity_score": "Opportunity Score",
        }
    )

    write_dataframe(ws, out, start_row=1)
    style_table_header(ws)
    style_data_cells(ws, 2, ws.max_row, ws.max_column)
    ws.freeze_panes = "A2"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "$#,##0.00"
        ws.cell(row=r, column=6).number_format = "$#,##0.00"
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = "0.0%"
        ws.cell(row=r, column=9).number_format = "0.0%"

        score_cell = ws.cell(row=r, column=10)
        score = score_cell.value
        if score >= 70:
            score_cell.fill = PatternFill(fill_type="solid", fgColor=DARK_GREEN_FILL)
            score_cell.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
        elif score >= 50:
            score_cell.fill = PatternFill(fill_type="solid", fgColor=MID_GREEN_FILL)
        elif score >= 30:
            score_cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
        else:
            score_cell.fill = PatternFill(fill_type="solid", fgColor=GRAY_FILL)

    auto_fit_columns(ws)


def build_capex_tab(wb: Workbook, properties_df: pd.DataFrame, capex_df: pd.DataFrame):
    ws = wb.create_sheet("CapEx Projects")
    ws.sheet_properties.tabColor = DARK_ORANGE

    capex = capex_df.copy()
    capex["start_date"] = pd.to_datetime(capex["start_date"])
    capex["end_date"] = pd.to_datetime(capex["end_date"])

    capex = capex.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )

    # KPI summary cards
    total_budgeted = float(capex["budgeted_cost"].sum())
    total_actual = float(capex["actual_cost"].sum())
    total_variance = float(capex["variance"].sum())
    projects_over_budget = int((capex["variance"] > 0).sum())

    kpi_defs = [
        ("Total Budgeted", total_budgeted, "$#,##0.00"),
        ("Total Actual", total_actual, "$#,##0.00"),
        ("Total Variance", total_variance, "$#,##0.00"),
        ("Projects Over Budget", projects_over_budget, "0"),
    ]

    starts = [1, 4, 7, 10]
    for i, (label, value, num_fmt) in enumerate(kpi_defs):
        start_col = starts[i]
        end_col = start_col + 2
        ws.merge_cells(start_row=1, start_column=start_col, end_row=2, end_column=end_col)
        ws.merge_cells(start_row=3, start_column=start_col, end_row=5, end_column=end_col)

        label_cell = ws.cell(row=1, column=start_col, value=label)
        value_cell = ws.cell(row=3, column=start_col, value=value)

        for r in range(1, 6):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
                cell.border = CARD_BORDER

        label_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_color = "FFFFFF"
        if label == "Total Variance":
            value_color = "FF4D4D" if total_variance > 0 else "92D050"
        value_cell.font = Font(name="Calibri", size=16, bold=True, color=value_color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = num_fmt

    detail = capex[
        [
            "property_name",
            "asset_type",
            "project_name",
            "project_type",
            "start_date",
            "end_date",
            "budgeted_cost",
            "actual_cost",
            "variance",
            "status",
        ]
    ].rename(
        columns={
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "project_name": "Project Name",
            "project_type": "Project Type",
            "start_date": "Start Date",
            "end_date": "End Date",
            "budgeted_cost": "Budgeted Cost",
            "actual_cost": "Actual Cost",
            "variance": "Variance",
            "status": "Status",
        }
    )

    data_start_row = 8
    write_dataframe(ws, detail, start_row=data_start_row)
    style_custom_header(ws, data_start_row, 1, detail.shape[1])
    style_data_cells(ws, data_start_row + 1, ws.max_row, ws.max_column)

    for r in range(data_start_row + 1, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "mmm-yy"
        ws.cell(row=r, column=6).number_format = "mmm-yy"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"
        ws.cell(row=r, column=8).number_format = "$#,##0.00"
        variance_cell = ws.cell(row=r, column=9)
        variance_cell.number_format = "$#,##0.00"
        variance_cell.font = Font(
            name="Calibri",
            size=10,
            color="C00000" if (variance_cell.value or 0) > 0 else "008000",
            bold=False,
        )

    ws.freeze_panes = "A9"
    auto_fit_columns(ws)


def build_dashboard_tab(
    wb: Workbook,
    properties_df: pd.DataFrame,
    occupancy_df: pd.DataFrame,
    delinquency_df: pd.DataFrame,
    financials_df: pd.DataFrame,
):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = RED_TAB

    occ = occupancy_df.copy()
    dlq = delinquency_df.copy()
    fin = financials_df.copy()

    occ["month"] = pd.to_datetime(occ["month"])
    dlq["month"] = pd.to_datetime(dlq["month"])
    fin["month"] = pd.to_datetime(fin["month"])
    occ["occupancy_rate"] = normalize_pct_series(occ["occupancy_rate"])
    dlq["delinquency_rate"] = normalize_pct_series(dlq["delinquency_rate"])
    fin["noi_margin"] = normalize_pct_series(fin["noi_margin"])
    fin["expense_ratio"] = normalize_pct_series(fin["expense_ratio"])

    latest_month = occ["month"].max()
    occ_latest = occ[occ["month"] == latest_month].copy()
    dlq_latest = dlq[dlq["month"] == latest_month].copy()
    fin_latest = fin[fin["month"] == latest_month].copy()

    occ_latest = occ_latest.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )
    dlq_latest = dlq_latest.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )
    fin_latest = fin_latest.merge(
        properties_df[["property_id", "property_name", "asset_type"]],
        on="property_id",
        how="left",
    )

    ws.merge_cells("A1:L1")
    ws["A1"] = "Executive Portfolio Dashboard"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = "Most Recent Month Performance Summary"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Section A - KPI cards
    total_properties = int(len(properties_df))
    total_units = int(properties_df["total_units"].sum())
    portfolio_avg_occ = occ_latest["occupied_units"].sum() / occ_latest["total_units"].sum()
    total_monthly_noi = float(fin_latest["noi"].sum())
    avg_delinquency = dlq_latest["past_due_amount"].sum() / dlq_latest["rent_billed"].sum()
    total_aum = float(properties_df["purchase_price"].sum())

    kpis = [
        ("Total Properties", total_properties, "0"),
        ("Total Units", total_units, "#,##0"),
        ("Portfolio Avg Occupancy", portfolio_avg_occ, "0.0%"),
        ("Total Monthly NOI", total_monthly_noi, "$#,##0.00"),
        ("Avg Delinquency Rate", avg_delinquency, "0.0%"),
        ("Total AUM", total_aum, "$#,##0.00"),
    ]

    card_spans = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    for (label, value, number_format), (start_col, end_col) in zip(kpis, card_spans):
        ws.merge_cells(start_row=3, start_column=start_col, end_row=4, end_column=end_col)
        ws.merge_cells(start_row=5, start_column=start_col, end_row=8, end_column=end_col)

        label_cell = ws.cell(row=3, column=start_col, value=label)
        value_cell = ws.cell(row=5, column=start_col, value=value)

        for r in range(3, 9):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
                cell.border = CARD_BORDER

        label_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        value_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.number_format = number_format

    # Section B - Asset type comparison
    ws.merge_cells("A10:C10")
    ws["A10"] = "Asset Type Comparison"
    ws["A10"].font = Font(name="Calibri", size=12, bold=True, color=DARK_NAVY)
    ws["A10"].alignment = Alignment(horizontal="left", vertical="center")

    metrics_header_row = 11
    comparison_rows = [
        "Avg Occupancy",
        "Avg Delinquency",
        "Total NOI",
        "Avg NOI Margin",
        "Avg Expense Ratio",
    ]

    mhc_occ = occ_latest[occ_latest["asset_type"] == "MHC"]
    stg_occ = occ_latest[occ_latest["asset_type"] == "Self-Storage"]
    mhc_dlq = dlq_latest[dlq_latest["asset_type"] == "MHC"]
    stg_dlq = dlq_latest[dlq_latest["asset_type"] == "Self-Storage"]
    mhc_fin = fin_latest[fin_latest["asset_type"] == "MHC"]
    stg_fin = fin_latest[fin_latest["asset_type"] == "Self-Storage"]

    comparison_values = [
        [
            mhc_occ["occupied_units"].sum() / mhc_occ["total_units"].sum(),
            stg_occ["occupied_units"].sum() / stg_occ["total_units"].sum(),
        ],
        [
            mhc_dlq["past_due_amount"].sum() / mhc_dlq["rent_billed"].sum(),
            stg_dlq["past_due_amount"].sum() / stg_dlq["rent_billed"].sum(),
        ],
        [
            mhc_fin["noi"].sum(),
            stg_fin["noi"].sum(),
        ],
        [
            mhc_fin["noi_margin"].mean(),
            stg_fin["noi_margin"].mean(),
        ],
        [
            mhc_fin["expense_ratio"].mean(),
            stg_fin["expense_ratio"].mean(),
        ],
    ]

    ws.cell(row=metrics_header_row, column=1, value="Metric")
    ws.cell(row=metrics_header_row, column=2, value="MHC")
    ws.cell(row=metrics_header_row, column=3, value="Self-Storage")
    style_custom_header(ws, metrics_header_row, 1, 3)

    for idx, metric in enumerate(comparison_rows, start=1):
        row = metrics_header_row + idx
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=comparison_values[idx - 1][0])
        ws.cell(row=row, column=3, value=comparison_values[idx - 1][1])

    style_data_cells(ws, metrics_header_row + 1, metrics_header_row + len(comparison_rows), 3)
    for r in range(metrics_header_row + 1, metrics_header_row + len(comparison_rows) + 1):
        metric_name = ws.cell(row=r, column=1).value
        if metric_name == "Total NOI":
            ws.cell(row=r, column=2).number_format = "$#,##0.00"
            ws.cell(row=r, column=3).number_format = "$#,##0.00"
        else:
            ws.cell(row=r, column=2).number_format = "0.0%"
            ws.cell(row=r, column=3).number_format = "0.0%"

    # Section C - Exception report
    ws.merge_cells("A20:G20")
    ws["A20"] = "Properties Requiring Attention — Most Recent Month"
    ws["A20"].font = Font(name="Calibri", size=12, bold=True, color=DARK_NAVY)
    ws["A20"].alignment = Alignment(horizontal="left", vertical="center")

    combined = (
        occ_latest[
            [
                "property_id",
                "property_name",
                "asset_type",
                "occupancy_rate",
            ]
        ]
        .merge(
            dlq_latest[
                ["property_id", "delinquency_rate"]
            ],
            on="property_id",
            how="inner",
        )
        .merge(
            fin_latest[
                ["property_id", "noi", "noi_margin"]
            ],
            on="property_id",
            how="inner",
        )
    )

    exceptions = combined[
        (combined["occupancy_rate"] < 0.90) | (combined["delinquency_rate"] > 0.08)
    ].copy()

    def risk_label(row):
        occ_risk = row["occupancy_rate"] < 0.90
        dlq_risk = row["delinquency_rate"] > 0.08
        if occ_risk and dlq_risk:
            return "HIGH RISK"
        if dlq_risk:
            return "DELINQUENCY RISK"
        if occ_risk:
            return "OCCUPANCY RISK"
        return "WATCH"

    priority = {"HIGH RISK": 1, "DELINQUENCY RISK": 2, "OCCUPANCY RISK": 3, "WATCH": 4}
    exceptions["Risk Label"] = exceptions.apply(risk_label, axis=1)
    exceptions["risk_order"] = exceptions["Risk Label"].map(priority)
    exceptions.sort_values(["risk_order", "delinquency_rate", "occupancy_rate"], ascending=[True, False, True], inplace=True)

    exception_out = exceptions[
        [
            "property_name",
            "asset_type",
            "occupancy_rate",
            "delinquency_rate",
            "noi",
            "noi_margin",
            "Risk Label",
        ]
    ].rename(
        columns={
            "property_name": "Property Name",
            "asset_type": "Asset Type",
            "occupancy_rate": "Occupancy Rate",
            "delinquency_rate": "Delinquency Rate",
            "noi": "NOI",
            "noi_margin": "NOI Margin",
        }
    )

    exception_header_row = 21
    write_dataframe(ws, exception_out, start_row=exception_header_row)
    style_custom_header(ws, exception_header_row, 1, 7)
    style_data_cells(ws, exception_header_row + 1, ws.max_row, 7)

    for r in range(exception_header_row + 1, exception_header_row + len(exception_out) + 1):
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=5).number_format = "$#,##0.00"
        ws.cell(row=r, column=6).number_format = "0.0%"

        label_cell = ws.cell(row=r, column=7)
        label = label_cell.value
        if label == "HIGH RISK":
            label_cell.fill = PatternFill(fill_type="solid", fgColor=RED_FILL)
            label_cell.font = Font(name="Calibri", size=10, bold=True, color=DARK_RED_TEXT)
        elif label == "DELINQUENCY RISK":
            label_cell.fill = PatternFill(fill_type="solid", fgColor=ORANGE_FILL)
            label_cell.font = Font(name="Calibri", size=10, bold=True)
        elif label == "OCCUPANCY RISK":
            label_cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
            label_cell.font = Font(name="Calibri", size=10, bold=True)
        else:
            label_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GRAY)
            label_cell.font = Font(name="Calibri", size=10)

    note_row = exception_header_row + len(exception_out) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    ws.cell(
        row=note_row,
        column=1,
        value="This exception report combines occupancy, collections risk, and NOI to help asset managers prioritize follow-up.",
    )
    ws.cell(row=note_row, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Section D - Top/Bottom 5 NOI Margin
    top_bottom_start = note_row + 2
    ws.merge_cells(start_row=top_bottom_start, start_column=1, end_row=top_bottom_start, end_column=3)
    ws.cell(row=top_bottom_start, column=1, value="Top 5 Properties by NOI Margin")
    ws.cell(row=top_bottom_start, column=1).font = Font(name="Calibri", size=12, bold=True, color=DARK_NAVY)

    ws.merge_cells(start_row=top_bottom_start, start_column=6, end_row=top_bottom_start, end_column=8)
    ws.cell(row=top_bottom_start, column=6, value="Bottom 5 Properties by NOI Margin")
    ws.cell(row=top_bottom_start, column=6).font = Font(name="Calibri", size=12, bold=True, color=DARK_NAVY)

    noi_rank = fin_latest[["property_id", "property_name", "asset_type", "noi_margin"]].copy()
    top5 = noi_rank.sort_values("noi_margin", ascending=False).head(5)
    bottom5 = noi_rank.sort_values("noi_margin", ascending=True).head(5)

    top_out = top5[["property_name", "asset_type", "noi_margin"]].rename(
        columns={"property_name": "Property Name", "asset_type": "Asset Type", "noi_margin": "NOI Margin"}
    )
    bottom_out = bottom5[["property_name", "asset_type", "noi_margin"]].rename(
        columns={"property_name": "Property Name", "asset_type": "Asset Type", "noi_margin": "NOI Margin"}
    )

    top_header = top_bottom_start + 1
    write_dataframe(ws, top_out, start_row=top_header, start_col=1)
    style_custom_header(ws, top_header, 1, 3)
    style_data_cells(ws, top_header + 1, top_header + len(top_out), 3)
    for r in range(top_header + 1, top_header + len(top_out) + 1):
        ws.cell(row=r, column=3).number_format = "0.0%"

    write_dataframe(ws, bottom_out, start_row=top_header, start_col=6)
    style_custom_header(ws, top_header, 6, 8)
    style_data_cells(ws, top_header + 1, top_header + len(bottom_out), 8)
    for r in range(top_header + 1, top_header + len(bottom_out) + 1):
        ws.cell(row=r, column=8).number_format = "0.0%"

    ws.freeze_panes = "A3"
    auto_fit_columns(ws)


def calculate_irr(cashflows, low=-0.95, high=1.5, tol=1e-7, max_iter=200):
    def npv(rate):
        return sum(cf / ((1 + rate) ** i) for i, cf in enumerate(cashflows))

    npv_low = npv(low)
    npv_high = npv(high)
    if npv_low * npv_high > 0:
        return 0.0

    for _ in range(max_iter):
        mid = (low + high) / 2
        npv_mid = npv(mid)
        if abs(npv_mid) < tol:
            return mid
        if npv_low * npv_mid < 0:
            high = mid
            npv_high = npv_mid
        else:
            low = mid
            npv_low = npv_mid
    return mid


def build_acq_model_tab(wb: Workbook):
    ws = wb.create_sheet("Acq. Model")
    ws.sheet_properties.tabColor = "FFD700"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Acquisition Underwriting — MHC-05 Pinon Pines"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # Assumptions
    ws.merge_cells("A2:C2")
    ws["A2"] = "Underwriting Assumptions"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    assumptions = [
        ("Purchase Price", 4200000, "$#,##0.00"),
        ("Total Units", 88, "0"),
        ("Current Occupancy", 0.82, "0.0%"),
        ("Stabilized Occupancy Target", 0.93, "0.0%"),
        ("Current Avg Monthly Rent per Lot", 720, "$#,##0.00"),
        ("Market Avg Rent", 810, "$#,##0.00"),
        ("Year 1 Rent Growth", 0.045, "0.0%"),
        ("Years 2-5 Rent Growth", 0.03, "0.0%"),
        ("Operating Expense Ratio Year 1", 0.48, "0.0%"),
        ("Expense Ratio Improvement", -0.01, "0.0%"),
        ("Renovation CapEx Year 1", 380000, "$#,##0.00"),
        ("Exit Cap Rate", 0.0625, "0.00%"),
        ("Loan Amount", 2730000, "$#,##0.00"),
        ("Interest Rate", 0.068, "0.0%"),
        ("Loan Term", 10, "0"),
        ("Acquisition Costs", 0.025, "0.0%"),
    ]

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for i, (label, value, fmt) in enumerate(assumptions, start=3):
        ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.number_format = fmt
        value_cell.fill = yellow_fill
        for c in (1, 2):
            ws.cell(row=i, column=c).font = Font(name="Calibri", size=10)
            ws.cell(row=i, column=c).border = THIN_BORDER

    ws.cell(row=3, column=1, value="Assumption").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=3, column=2, value="Value").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=3, column=1).fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    ws.cell(row=3, column=2).fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    ws.cell(row=3, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=3, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Model calculations
    purchase_price = 4200000
    total_units = 88
    current_occupancy = 0.82
    stabilized_occupancy = 0.93
    current_rent = 720
    year1_growth = 0.045
    years_2_5_growth = 0.03
    expense_ratio_y1 = 0.48
    expense_improvement = -0.01
    renovation_capex = 380000
    exit_cap_rate = 0.0625
    loan_amount = 2730000
    interest_rate = 0.068
    loan_term = 10
    acquisition_cost_pct = 0.025
    acquisition_costs = purchase_price * acquisition_cost_pct
    initial_equity = purchase_price - loan_amount + renovation_capex + acquisition_costs

    occupancy_rates = [0.84, 0.865, 0.89, 0.91, stabilized_occupancy]
    avg_monthly_rents = [current_rent * (1 + year1_growth)]
    for _ in range(4):
        avg_monthly_rents.append(avg_monthly_rents[-1] * (1 + years_2_5_growth))
    expense_ratios = [expense_ratio_y1 + (i * expense_improvement) for i in range(5)]

    gpr = [total_units * r * 12 for r in avg_monthly_rents]
    vacancy_loss = [gpr[i] * (1 - occupancy_rates[i]) for i in range(5)]
    egi = [gpr[i] - vacancy_loss[i] for i in range(5)]
    op_ex = [egi[i] * expense_ratios[i] for i in range(5)]
    noi = [egi[i] - op_ex[i] for i in range(5)]
    noi_margin = [noi[i] / egi[i] if egi[i] else 0 for i in range(5)]

    # Annual debt service payment (P&I)
    annual_rate = interest_rate
    n_periods = loan_term
    annual_debt_service = loan_amount * annual_rate / (1 - (1 + annual_rate) ** (-n_periods))
    debt_service = [annual_debt_service] * 5

    net_cash_flow = [noi[i] - debt_service[i] for i in range(5)]
    coc_return = [net_cash_flow[i] / initial_equity for i in range(5)]
    cumulative_cash_flow = []
    running = 0
    for val in net_cash_flow:
        running += val
        cumulative_cash_flow.append(running)

    # Pro forma layout
    ws.merge_cells("A22:F22")
    ws["A22"] = "5-Year Operating Pro Forma"
    ws["A22"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A22"].fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
    ws["A22"].alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=23, column=1, value="Metric")
    for i in range(1, 6):
        ws.cell(row=23, column=i + 1, value=f"Year {i}")
    style_custom_header(ws, 23, 1, 6)

    metrics = [
        ("Occupied Units", [round(total_units * r, 0) for r in occupancy_rates], "0"),
        ("Occupancy Rate", occupancy_rates, "0.0%"),
        ("Avg Monthly Rent", avg_monthly_rents, "$#,##0.00"),
        ("Gross Potential Rent Annual", gpr, "$#,##0.00"),
        ("Vacancy Loss", vacancy_loss, "$#,##0.00"),
        ("Effective Gross Income", egi, "$#,##0.00"),
        ("Operating Expenses", op_ex, "$#,##0.00"),
        ("NOI", noi, "$#,##0.00"),
        ("NOI Margin %", noi_margin, "0.0%"),
        ("Debt Service Annual P&I", debt_service, "$#,##0.00"),
        ("Net Cash Flow After Debt", net_cash_flow, "$#,##0.00"),
        ("Cash-on-Cash Return", coc_return, "0.0%"),
        ("Cumulative Cash Flow", cumulative_cash_flow, "$#,##0.00"),
    ]

    for row_offset, (metric_name, values, fmt) in enumerate(metrics, start=24):
        ws.cell(row=row_offset, column=1, value=metric_name)
        ws.cell(row=row_offset, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=row_offset, column=1).border = THIN_BORDER
        for i, value in enumerate(values, start=2):
            cell = ws.cell(row=row_offset, column=i, value=float(value))
            cell.number_format = fmt
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER

    # Exit analysis
    ws.merge_cells("A39:D39")
    ws["A39"] = "Exit Analysis"
    ws["A39"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A39"].fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)

    years_paid = 5
    remaining_balance = (
        loan_amount * ((1 + annual_rate) ** years_paid)
        - annual_debt_service * ((((1 + annual_rate) ** years_paid) - 1) / annual_rate)
    )
    year5_exit_value = noi[4] / exit_cap_rate
    net_proceeds = year5_exit_value - remaining_balance
    total_return = net_proceeds + cumulative_cash_flow[4]
    equity_multiple = total_return / initial_equity

    cashflows = [-initial_equity] + net_cash_flow[:4] + [net_cash_flow[4] + net_proceeds]
    irr = calculate_irr(cashflows)

    exit_rows = [
        ("Year 5 Exit Value", year5_exit_value, "$#,##0.00"),
        ("Remaining Loan Balance after 5 years", remaining_balance, "$#,##0.00"),
        ("Net Proceeds", net_proceeds, "$#,##0.00"),
        ("Cumulative Cash Flow", cumulative_cash_flow[4], "$#,##0.00"),
        ("Total Return", total_return, "$#,##0.00"),
        ("Initial Equity", initial_equity, "$#,##0.00"),
        ("Equity Multiple", equity_multiple, "0.00x"),
        ("IRR", irr, "0.0%"),
    ]

    for idx, (label, value, fmt) in enumerate(exit_rows, start=40):
        ws.cell(row=idx, column=1, value=label)
        val_cell = ws.cell(row=idx, column=2, value=float(value))
        val_cell.number_format = fmt
        for c in (1, 2):
            ws.cell(row=idx, column=c).font = Font(name="Calibri", size=10)
            ws.cell(row=idx, column=c).border = THIN_BORDER

    # Sensitivity table
    ws.merge_cells("A50:F50")
    ws["A50"] = "IRR Sensitivity — Exit Cap Rate vs Year 5 NOI Scenario"
    ws["A50"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws["A50"].fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)

    cap_rates = [0.055, 0.0575, 0.06, 0.0625, 0.065, 0.0675, 0.07]
    noi_scenarios = [("−5%", 0.95), ("0%", 1.00), ("Base", 1.00), ("+5%", 1.05), ("+10%", 1.10)]

    ws.cell(row=51, column=1, value="Exit Cap Rate")
    for j, (label, _) in enumerate(noi_scenarios, start=2):
        ws.cell(row=51, column=j, value=label)
    style_custom_header(ws, 51, 1, 6)

    for i, cap in enumerate(cap_rates, start=52):
        ws.cell(row=i, column=1, value=cap)
        ws.cell(row=i, column=1).number_format = "0.00%"
        ws.cell(row=i, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=1).border = THIN_BORDER

        for j, (_, mult) in enumerate(noi_scenarios, start=2):
            adjusted_noi = noi[4] * mult
            exit_value = adjusted_noi / cap
            net_proceeds_sens = exit_value - remaining_balance
            final_year_cf = net_cash_flow[4] + net_proceeds_sens
            irr_sens = calculate_irr([-initial_equity] + net_cash_flow[:4] + [final_year_cf])

            cell = ws.cell(row=i, column=j, value=float(irr_sens))
            cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            if irr_sens > 0.15:
                cell.fill = PatternFill(fill_type="solid", fgColor=GREEN_FILL)
            elif irr_sens >= 0.10:
                cell.fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL)
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor=RED_FILL)

    auto_fit_columns(ws)


def build_insights_tab(wb: Workbook):
    ws = wb.create_sheet("Insights")
    ws.sheet_properties.tabColor = "006100"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Portfolio Findings & Recommended Actions"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "Based on 12-month data analysis, January–December 2024"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    findings = [
        (
            "STG-02: Collections Problem, Not a Demand Problem",
            "SunState Storage | Self-Storage | Good occupancy but delinquency rising from approximately 7% to 13%",
            "STG-02 maintained generally healthy occupancy, which suggests demand is not the primary issue. The bigger concern is the upward delinquency trend, meaning revenue leakage is coming from collections rather than lack of tenants.",
            "Review aged receivables, late-fee enforcement, payment reminders, and property-level collection workflows. The asset management team should work with property operations to reduce delinquency before it impacts NOI further.",
        ),
        (
            "MHC-03: Dual-Flag Asset Requiring Immediate Operational Review",
            "Canyon Ridge | MHC | Low occupancy around 78–83% and high delinquency around 12–16%",
            "MHC-03 is the weakest risk profile in the portfolio because it has both demand weakness and collections risk. This combination can pressure revenue, NOI, and investor reporting.",
            "Prioritize an operational review covering leasing activity, local market positioning, rent levels, resident retention, and collections. This property should be included in weekly asset management follow-up until both occupancy and delinquency stabilize.",
        ),
        (
            "MHC-04: Portfolio Benchmark — Operational Best Practices Template",
            "Saguaro Estates | MHC | High occupancy, low delinquency, strong NOI margin",
            "MHC-04 performs like a benchmark asset, with high occupancy, low delinquency, and strong profitability. This indicates strong property-level execution and stable resident demand.",
            "Use MHC-04 as an internal operating benchmark. Compare staffing, collections process, leasing practices, resident communication, and expense controls against weaker MHC assets.",
        ),
        (
            "MHC-05: Value-Add Play — Track Against Pro Forma Monthly",
            "Pinon Pines | MHC | Occupancy improves from roughly 81% to 89% over the year",
            "MHC-05 shows a positive value-add trend with improving occupancy, but it has not yet reached stabilized performance. The asset has rent upside and operational upside, making it a good candidate for active tracking.",
            "Track MHC-05 monthly against the acquisition pro forma, especially occupancy, rent growth, CapEx spend, and NOI margin. Management should confirm that renovation dollars are converting into occupancy gains and stronger cash flow.",
        ),
        (
            "STG-03: Expense Ratio Drag — NOI Underperformance Despite Adequate Occupancy",
            "Desert Vault | Self-Storage | Expense ratio around 62–68%",
            "STG-03 does not appear to be purely an occupancy problem. The high expense ratio is reducing NOI margin and making the property less profitable than it should be.",
            "Review controllable expenses such as repairs, payroll, utilities, vendor contracts, insurance, and recurring maintenance. The finance and asset management teams should isolate which expense categories are driving margin compression.",
        ),
        (
            "MHC-07: Early Warning — Delinquency Trend Requires Collections Intervention Before It Escalates",
            "Mesa Vista | MHC | High occupancy but delinquency rising from approximately 5% to 11%",
            "MHC-07 still has strong occupancy, but the delinquency trend is moving in the wrong direction. This is an early warning sign because the property looks healthy from an occupancy view but has growing collection risk.",
            "Intervene before the issue becomes critical by reviewing tenant aging buckets, payment plan usage, late-fee timing, and property manager follow-up. This should be treated as a collections process issue before assuming a demand issue.",
        ),
    ]

    row = 4
    for title, metric, shows, action in findings:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=10)
        ws.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=10)
        ws.merge_cells(start_row=row + 3, start_column=1, end_row=row + 3, end_column=10)

        title_cell = ws.cell(row=row, column=1, value=title)
        metric_cell = ws.cell(row=row + 1, column=1, value=metric)
        data_cell = ws.cell(row=row + 2, column=1, value=f"Data shows: {shows}")
        action_cell = ws.cell(row=row + 3, column=1, value=f"Recommended action: {action}")

        title_cell.fill = PatternFill(fill_type="solid", fgColor=DARK_NAVY)
        title_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        metric_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_BLUE)
        metric_cell.font = Font(name="Calibri", size=10, bold=True, color="000000")

        data_cell.fill = PatternFill(fill_type="solid", fgColor=WHITE)
        data_cell.font = Font(name="Calibri", size=10)

        action_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GREEN)
        action_cell.font = Font(name="Calibri", size=10, italic=True)

        for r in range(row, row + 4):
            cell = ws.cell(row=r, column=1)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = THIN_BORDER

        ws.row_dimensions[row].height = 28
        ws.row_dimensions[row + 1].height = 30
        ws.row_dimensions[row + 2].height = 55
        ws.row_dimensions[row + 3].height = 55
        row += 5

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 20


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    properties_df = pd.read_csv(DATA_DIR / "properties.csv")
    occupancy_df = pd.read_csv(DATA_DIR / "monthly_occupancy.csv")
    delinquency_df = pd.read_csv(DATA_DIR / "delinquency.csv")
    financials_df = pd.read_csv(DATA_DIR / "property_financials.csv")
    market_df = pd.read_csv(DATA_DIR / "market_comps.csv")
    capex_df = pd.read_csv(DATA_DIR / "capex_projects.csv")

    wb = Workbook()
    wb.remove(wb.active)

    build_cover_tab(wb)
    build_properties_tab(wb, properties_df)
    build_occupancy_tab(wb, properties_df, occupancy_df)
    build_delinquency_tab(wb, properties_df, delinquency_df)
    build_financials_tab(wb, properties_df, financials_df)
    build_market_comps_tab(wb, properties_df, market_df)
    build_capex_tab(wb, properties_df, capex_df)
    build_dashboard_tab(wb, properties_df, occupancy_df, delinquency_df, financials_df)
    build_acq_model_tab(wb)
    build_insights_tab(wb)

    wb.save(OUTPUT_PATH)

    print("Final workbook complete with 10 tabs: excel/portfolio_dashboard.xlsx")
    print(f"Sheets created: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
